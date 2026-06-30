from __future__ import annotations
import os
from datetime import datetime
from pathlib import Path
from typing import Any
from langchain_core.tools import StructuredTool
from react_agent.utils.tool_helpers import _ok, _err, ExcelInput
from react_agent.core.config import settings


# ================================================================================
# 工具 2：Excel 表格生成
# ================================================================================
# 安装依赖：pip install openpyxl
# 配置环境变量：
#   - EXCEL_OUTPUT_DIR：Excel 文件输出目录（默认 ./outputs）
#   - RAG_MAX_CONTENT_CHARS：内容字符数上限（默认 800）
async def _make_excel_impl(
        filename: str,
        headers: list[str],
        rows: list[list[Any]],
        sheet_name: str = "Sheet1",
        mode: str = settings.tools.excel.mode,     # "timestamp" | "overwrite" | "append"
        keep_backup: bool = settings.tools.excel.keep_backup,   # append/overwrite 时是否额外留一份时间戳备份
) -> dict[str, Any]:
    """
    Excel 表格生成的核心实现。

    功能描述：
      - 创建或修改 Excel 工作簿
      - 支持三种模式：timestamp（带时间戳）、overwrite（覆盖）、append（追加）
      - 自动美化表头：粗体、居中、灰色背景
      - 自动调整列宽以适应内容
      - 冻结表头行便于滚动查看

    参数：
      filename (str)：不带后缀的文件名（会自动补 .xlsx）
      headers (list[str])：表头，不能为空
      rows (list[list[Any]])：数据行列表
      sheet_name (str)：工作表名称（默认 "Sheet1"，不能超过 31 字符）
      mode (str)：模式选择
                  - "timestamp"：添加时间戳前缀，新建文件，不覆盖旧数据
                  - "overwrite"：覆盖已有的同名文件，危险操作
                  - "append"：追加到已有文件，要求表头一致
      keep_backup (bool)：在 append/overwrite 时是否保存时间戳备份

    执行流程：
      (1) 输入校验：filename/headers/rows 不能为空或格式错误
      (2) Mode 校验：mode 必须是允许值之一
      (3) 依赖检查：openpyxl 库
      (4) 输出目录创建：EXCEL_OUTPUT_DIR 环境变量，默认 ./outputs
      (5) 模式分支：
          - timestamp 或新建：创建新 Workbook，写入表头和样式
          - append：打开已有文件，检查表头一致性，找到末行后追加
      (6) 追加数据行：遍历 rows，每行用 ws.append()
      (7) 列宽调整：扫描前 200 行，计算最大宽度，设置列宽
      (8) 可选备份：append/overwrite 时保存 .backup_timestamp 文件
      (9) 保存并返回：返回保存路径和元数据

    错误处理：
      - BAD_INPUT：参数缺失或无效
      - MISSING_DEPENDENCY：缺少 openpyxl
      - HEADER_MISMATCH：append 时表头不一致
      - EXCEL_PERMISSION_DENIED：文件被占用或无写权限
      - EXCEL_CREATE_FAILED：其他创建错误

    示例：
      result = await _make_excel_impl(
          filename="report",
          headers=["姓名", "部门"],
          rows=[["张三", "技术"], ["李四", "销售"]],
          mode="timestamp"
      )
      # 返回：
      # {
      #   "ok": True,
      #   "data": {
      #     "path": "/path/to/report_20260310_144200.xlsx",
      #     "base_path": "/path/to/report.xlsx",
      #     "sheet": "Sheet1",
      #     "row_count": 2,
      #     "mode": "timestamp"
      #   }
      # }
    """
    tool_name = "make_excel_table"
    fn = (filename or "").strip()

    # ════════════════════ 阶段 1：输入校验 ════════════════════
    if not fn:
        return _err(tool_name=tool_name, query="", code="BAD_INPUT", message="filename 不能为空")
    if not fn.lower().endswith(".xlsx"):
        fn += ".xlsx"

    if not headers or not isinstance(headers, list):
        return _err(tool_name=tool_name, query=fn, code="BAD_INPUT", message="headers 必须是非空 list[str]")
    if rows is None or not isinstance(rows, list):
        return _err(tool_name=tool_name, query=fn, code="BAD_INPUT", message="rows 必须是 list[list[Any]]")

    # ════════════════════ 阶段 2：Mode 校验 ════════════════════
    valid_modes = {"timestamp", "overwrite", "append"}
    if mode not in valid_modes:
        return _err(
            tool_name=tool_name, query=fn, code="BAD_INPUT",
            message=f"mode 参数无效，必须是 {sorted(valid_modes)} 之一，当前值：{mode!r}"
        )

    # ════════════════════ 阶段 3：依赖检查 ════════════════════
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return _err(
            tool_name=tool_name,
            query=fn,
            code="MISSING_DEPENDENCY",
            message="缺少依赖 openpyxl。请安装：pip install openpyxl"
        )

    # ════════════════════ 阶段 4：输出目录准备 ════════════════════
    out_dir = os.getenv("OUTPUT_DIR", "./outputs")
    out_dir = Path(out_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    base_path = out_dir / fn

    # ════════════════════ 阶段 5a：确定输出路径 ════════════════════
    # 根据 mode 决定是否添加时间戳前缀
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if mode == "timestamp":
        # 新建带时间戳的文件，避免覆盖已有文件
        out_path = out_dir / f"{Path(fn).stem}_{stamp}.xlsx"
    else:
        # 覆盖或追加到固定路径
        out_path = base_path

    # ════════════════════ 阶段 5b：创建或打开工作簿 ════════════════════
    wb = None
    ws = None

    if mode == "append" and base_path.exists():
        # ════ 追加模式：打开已有文件 ════
        wb = load_workbook(base_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

        # 检查表头一致性（防止列错位）
        existing_headers = [ws.cell(row=1, column=i).value for i in range(1, len(headers) + 1)]

        if any(h is None for h in existing_headers):
            # 表头行为空（可能是新表），原地写入表头
            # 注意：不能用 ws.append([...])，因为那会追加到末尾，而不是写入第一行
            for col_idx, h in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=str(h))
        elif [str(x) for x in existing_headers] != [str(h) for h in headers]:
            # 表头不一致，拒绝追加（保护数据）
            return _err(
                tool_name=tool_name,
                query=str(base_path),
                code="HEADER_MISMATCH",
                message="追加失败：目标文件表头与本次 headers 不一致（为防止列错位已拒绝写入）",
                meta={"existing": existing_headers, "incoming": headers},
            )
    else:
        # ════ 创建新表 (timestamp/overwrite) ════
        wb = Workbook()
        ws = wb.active
        ws.title = (sheet_name or "Sheet1")[:31]  # Excel 工作表名不超过 31 字符
        ws.append([str(h) for h in headers])

        # 美化表头：粗体、居中、灰色背景
        header_font = Font(bold=True)
        header_align = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill("solid", fgColor="EDEDED")  # 灰色
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = header_align
            cell.fill = header_fill

        # 冻结表头行，便于向下滚动时仍能看到列名
        ws.freeze_panes = "A2"

    # ════════════════════ 阶段 6：追加数据行 ════════════════════
    # 无论是新建还是追加，都在这里写入数据行
    for r in rows:
        ws.append(list(r))

    # ════════════════════ 阶段 7：自动调整列宽 ════════════════════
    # 扫描前 200 行内容，计算最大宽度，自动设置列宽（提升用户体验）
    max_width_cap = 60  # 列宽上限，防止过宽
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        # 初始宽度：表头的长度
        max_len = len(str(headers[col_idx - 1] or ""))
        # 扫描数据行（只扫前 200 行以避免性能问题）
        scan_n = min(len(rows), 200)
        for i in range(scan_n):
            try:
                v = rows[i][col_idx - 1]
            except Exception:
                v = ""
            max_len = max(max_len, len(str(v or "")))
        # 设置列宽：最多 60，最少 10，加 2 作为边距
        ws.column_dimensions[col_letter].width = min(max_width_cap, max(10, max_len + 2))

    # ════════════════════ 阶段 8：可选备份 ════════════════════
    # 在 append/overwrite 时可选地保存时间戳备份，便于数据恢复
    if keep_backup and base_path.exists() and mode in {"append", "overwrite"}:
        backup_path = out_dir / f"{Path(fn).stem}_backup_{stamp}.xlsx"
        # 用二进制复制，快速且安全
        backup_path.write_bytes(base_path.read_bytes())

    # ════════════════════ 阶段 9：保存文件并返回 ════════════════════
    try:
        wb.save(out_path)
        return _ok(
            tool_name=tool_name,
            query=fn,
            data={
                "path": str(out_path),          # 实际保存路径（可能带时间戳）
                "base_path": str(base_path),    # 基础路径（无时间戳）
                "sheet": ws.title,              # 工作表名
                "row_count": len(rows),         # 数据行数（不含表头）
                "mode": mode
            },
            meta={"output_dir": str(out_dir)},
        )
    except PermissionError as e:
        # 权限错误（文件被占用、无写权限）
        return _err(
            tool_name=tool_name,
            query=str(out_path),
            code="EXCEL_PERMISSION_DENIED",
            message=f"Excel 保存失败（权限/被占用）：{e}",
        )
    except Exception as e:
        # 其他错误（磁盘满、格式错误等）
        return _err(
            tool_name=tool_name,
            query=str(out_path),
            code="EXCEL_CREATE_FAILED",
            message=f"Excel 创建失败：{type(e).__name__}: {e}",
        )


# 用 LangChain StructuredTool 包装 _make_excel_impl
# StructuredTool 会根据 args_schema 生成工具的 JSON Schema，供 LLM 调用
make_excel_table = StructuredTool.from_function(
    coroutine=_make_excel_impl,   # 异步函数用 coroutine= 参数
    name="make_excel_table",
    description=(
        "【触发条件】以下情况调用本工具，将结构化数据保存为 Excel 文件：\n"
        "  - 用户要求导出、下载、保存表格数据\n"
        "  - 需要对比多公司/多指标数据，且数据条数 ≥ 3 行\n"
        "  - 用户明确说'生成Excel'、'整理成表格'、'保存数据'\n"
        "  示例场景：'把这5家公司的毛利率对比保存成Excel'\n\n"
        "【不触发条件】以下情况不调用本工具：\n"
        "  - 用户只需要在对话中看表格，未要求保存文件\n"
        "  - 数据仅有1-2行，直接回答即可\n\n"
        "【mode 选择逻辑】\n"
        "  timestamp（默认）— 每次新建带时间戳的文件，不覆盖历史，推荐首次使用\n"
        "  append           — 向已有同名文件追加数据，要求表头完全一致\n"
        "  overwrite        — 覆盖已有文件，数据会丢失，需用户明确要求时才用\n\n"
        "【输入关键字段】\n"
        "  filename  — 不带后缀的文件名，如 'byd_analysis'\n"
        "  headers   — 列名列表，如 ['公司', '毛利率', '报告期']\n"
        "  rows      — 数据行列表，每行与 headers 等长\n\n"
        "【输出关键字段】\n"
        "  data.path      — 实际保存的文件路径（含时间戳），告知用户此路径\n"
        "  data.row_count — 写入的数据行数（不含表头）"
    ),
    args_schema=ExcelInput,
)