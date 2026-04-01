# SNAPSHOT: (2026-01-31-14:26)
"""
================================================================================
                          LangGraph 工具层核心实现
================================================================================

目标：提供”工程可用”的工具层（对 Prompt 友好、对调试友好、对模型友好）
- 明确的工具 schema（@tool）：所有工具都有清晰的参数和描述
- 统一返回结构（ok/data/error/meta）：便于前端解析，便于调试
- 参数校验 + 异常兜底：输入检查 + 错误处理 + 超时重试
- 限制返回体积：避免上下文污染，裁剪过大返回

工具列表：
  1. search：Tavily Web 搜索（获取最新信息、验证事实）
  2. make_excel_table：Excel 表格创建/追加/覆盖（数据导出）
  3. query_internal_knowledge：RAG 私有知识库检索（内部文档、SOP、技术方案）

核心特性：
  - 超时控制 + 指数退避重试（with_retry 装饰器）
  - 线程安全的单例初始化（双重检查锁定）
  - 异步/同步双支持（async/await 兼容）
  - 模型缓存（BGE 模型、Reranker 等）
================================================================================
"""
from __future__ import annotations
import os
import time
import asyncio
import random
import inspect
import functools
import threading
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, List, Optional, cast, Dict, Type, Iterable
from langgraph.runtime import get_runtime
from langchain_core.tools import tool, StructuredTool
from pydantic import BaseModel, Field

from react_agent.memory.context import Context


# ================================================================================
# 统一返回结构：所有工具都返回这个结构，便于上游统一处理
# ================================================================================
def _ok(*, tool_name: str, query: str, data: Any, meta: Optional[Dict[str, Any]] = None):
    """
    返回成功结果。

    参数：
      tool_name (str)：工具名称，用于日志和追踪
      query (str)：原始查询字符串，便于问题定位
      data (Any)：实际返回的数据（search 返回结果、Excel 保存路径等）
      meta (dict)：元数据，如重试次数、耗时、是否缓存等（可选）

    返回：
      {
        “ok”: True,                 # 标志成功
        “tool”: tool_name,          # 工具标识
        “query”: query,             # 原始查询
        “data”: data,               # 核心数据
        “error”: None,              # 成功时无错误
        “meta”: {...}               # 执行元数据
      }
    """
    return {
        "ok": True,
        "tool": tool_name,
        "query": query,
        "data": data,
        "error": None,
        "meta": meta or {}
    }


def _err(*, tool_name: str, query: str, message: str, code: str = "TOOL_ERROR", meta: Optional[Dict[str, Any]] = None):
    """
    返回失败结果。

    参数：
      tool_name (str)：工具名称
      query (str)：原始查询字符串
      message (str)：人类可读的错误描述
      code (str)：错误代码（如 'BAD_INPUT', 'MISSING_DEPENDENCY', 'TIMEOUT'）
      meta (dict)：元数据，如已尝试次数等（可选）

    返回：
      {
        “ok”: False,                # 标志失败
        “tool”: tool_name,
        “query”: query,
        “data”: None,               # 失败时无数据
        “error”: {
          “code”: code,             # 错误类型代码
          “message”: message        # 错误详细说明
        },
        “meta”: {...}
      }
    """
    return {
        "ok": False,
        "tool": tool_name,
        "query": query,
        "data": None,
        "error": {"code": code, "message": message},
        "meta": meta or {}
    }

def _trim_text(s: str, max_chars: int):
    """
    字符串裁剪工具。

    功能：将过长的字符串截断，并在末尾添加”...”作为截断标记。
    用途：限制搜索结果、知识库内容等的大小，避免超长文本污染上下文。

    参数：
      s (str)：原始字符串，会先做 strip() 处理
      max_chars (int)：最大允许的字符数

    返回：
      - 如果 len(s) <= max_chars：返回原字符串（strip 后）
      - 否则：返回截断后的字符串 + “...” 作为截断标记

    示例：
      _trim_text(“Hello World”, 5) → “He...”  # 截断保留 5 字符
      _trim_text(“Hi”, 5) → “Hi”              # 不足 5 字符，原样返回
    """
    s = (s or "").strip()
    if len(s) <= max_chars:
        return s
    return s[: max_chars - 3] + "..."  # 预留 3 位空间给”...”


def _shrink_search_results(raw: Any, *, max_items: int = 5, max_char_per_item: int = 800):
    """
    搜索结果裁剪器。

    功能：把 Tavily API 返回的大块数据压缩到适合放入对话上下文的大小。
    用途：Web 搜索工具的后处理，过滤和裁剪结果以保持低成本。

    处理逻辑：
      1. 处理 None 输入 → 返回 None
      2. 处理标准 Tavily 格式 {“results”: [...], “answer”: “...”}
         - 取前 max_items 条结果
         - 每条结果裁剪标题（200 字）、内容（max_char_per_item 字）
         - 保留 score 和 published_date 作为参考
      3. 处理字符串格式 → 直接裁剪到 3000 字
      4. 其他格式 → 原样返回

    参数：
      raw (Any)：Tavily 原始返回或其他格式
      max_items (int)：最多保留几条搜索结果（默认 5）
      max_char_per_item (int)：每条结果内容的最大字符数（默认 800）

    返回：
      {
        “results”: [
          {
            “title”: str,            # 搜索结果标题（最多 200 字）
            “url”: str,              # 来源链接
            “content”: str,          # 摘要文本（最多 max_char_per_item 字）
            “score”: float,          # 匹配分数（可能为 None）
            “published_date”: str    # 发布日期（可能为 None）
          },
          ...
        ],
        “answer”: str                # Tavily 直接生成的简洁回答（最多 1200 字）
      }

    兼容性：
      - 不同版本 Tavily 可能返回不同字段，此函数使用 .get() 做容错处理
    """
    if raw is None:
        return None

    # 常见结构：{“results”:[{“title”:..., “url”:..., “content”:...}, ...], ...}
    if isinstance(raw, dict) and "results" in raw and isinstance(raw["results"], list):
        results = []
        for item in raw["results"][:max_items]:
            if not isinstance(item, dict):
                continue
            results.append(
                {
                    "title": _trim_text(str(item.get("title", "")), 200),
                    "url": str(item.get("url", "")),
                    "content": _trim_text(str(item.get("content", "")), max_char_per_item),
                    # 有些版本还有 score / published_date 等字段，不强依赖
                    "score": item.get("score"),
                    "published_date": item.get("published_date") or item.get("published_time"),
                }
            )
        # 保留少量 meta（避免太大）
        return {
            "results": results,
            "answer": _trim_text(str(raw.get("answer", "")), 1200),  # raw 在此分支必为 dict
        }

    # 如果不是预期结构，原样返回，但做一次字符串裁剪（保险）
    if isinstance(raw, str):
        return _trim_text(raw, 3000)

    return raw


# ================================================================================
# 通用工具装饰器：超时 + 指数退避重试 + 异步/同步双支持
# ================================================================================
def with_retry(
        *,
        tool_name: str,
        max_retries: int = 3,
        timeout: float = 10.0,
        base_delay: float = 0.5,
        max_delay: float = 6.0,
        retry_on: Iterable[Type[BaseException]] = (TimeoutError, ConnectionError, OSError)
):
    """
    为工具函数增强：超时控制 + 智能重试 + 指数退避策略。

    核心功能：
      1. 超时保护：如果工具调用超过 timeout 秒，立即中止
      2. 智能重试：对指定的异常类型（如 TimeoutError、ConnectionError）进行重试
      3. 指数退避：重试间隔逐次增加（0.5s → 1s → 2s → ...），避免冲击被调用方
      4. 异步/同步双支持：自动检测被装饰函数是异步还是同步
      5. 事件循环兼容性：处理 "已有事件循环" 场景（LangGraph/FastAPI 中）
      6. 统一返回结构：失败也返回 {ok: False, error: {...}, meta: {...}}

    参数：
      tool_name (str)：工具名称，用于错误日志和返回结构
      max_retries (int)：最多重试几次（默认 3 次，即最多执行 4 次）
      timeout (float)：单次调用超时时间，单位秒（默认 10 秒）
      base_delay (float)：首次重试延迟基数，单位秒（默认 0.5 秒）
      max_delay (float)：重试延迟上限，单位秒（默认 6 秒，防止等待时间过长）
      retry_on (Iterable)：哪些异常类型触发重试
                           （默认：TimeoutError, ConnectionError, OSError）
                           注意：工具函数需要主动 raise 这些异常，
                           不能在内部 try-except 中吞掉

    执行流程：
      (1) 第 1 次尝试（attempt=0）：
          - 调用工具函数，加 timeout 限制
          - 如果成功返回：记录 attempt/timeout/elapsed 到 meta，返回结果
          - 如果异常：捕获异常，判断是否需要重试

      (2) 异常判断：
          - 如果异常在 retry_on 列表中：进入重试流程
          - 否则：立即返回失败（不再尝试），因为非临时错误

      (3) 重试延迟：
          - 计算延迟时间：min(max_delay, base_delay * 2^attempt) * 随机偏移 (0.8~1.2)
          - 等待后继续第 2 次尝试

      (4) 重复 (1)~(3)，最多执行 max_retries + 1 次

      (5) 全部失败：返回 {ok: False, error: {...}, meta: {retries: 实际尝试次数}}

    使用示例：
      @with_retry(tool_name="my_api", max_retries=2, timeout=5)
      async def fetch_data(url: str) -> dict:
          # 函数内容
          # 关键：对于超时/连接错误，需要主动 raise，不能吞掉！
          try:
              result = await asyncio.wait_for(client.get(url), timeout=3)
          except asyncio.TimeoutError:
              raise  # 让 with_retry 捕获并重试

    注意事项：
      - 工具必须返回 dict，且应包含 "meta" 字段
      - 对临时错误（超时、连接拒绝）需要主动 raise，不能在工具内处理
      - 对永久错误（参数错误、依赖缺失）不应该重试，直接返回 _err(...)
      - 指数退避可以避免多个重试工具同时冲击后端，提高成功率
    """
    def decorator(fn):
        is_async = inspect.iscoroutinefunction(fn)

        # 预计算，避免每次循环都重建 tuple（性能优化）
        _retry_on_tuple = tuple(retry_on) if retry_on else ()

        async def _call_async(*args, **kwargs):
            """
            异步执行函数，带重试和超时保护。
            """
            last_exc: BaseException | None = None
            q = ""
            # 尽量从 kwargs/args 提取 query 方便定位问题所在
            if "query" in kwargs:
                q = str(kwargs.get("query") or "").strip()
            elif args:
                q = str(args[0] or "").strip()

            # 重试循环
            for attempt in range(max_retries + 1):
                try:
                    start = time.time()
                    # 异步和同步函数的调用方式不同
                    if is_async:
                        # 异步函数：直接 await，加 timeout 限制
                        res = await asyncio.wait_for(fn(*args, **kwargs), timeout=timeout)
                    else:
                        # 同步函数：用 asyncio.to_thread 在线程池中执行，避免阻塞事件循环
                        res = await asyncio.wait_for(asyncio.to_thread(fn, *args, **kwargs), timeout=timeout)

                    # 返回值检查：必须是 dict，且包含 meta 字段
                    if not isinstance(res, dict):
                        return _err(
                            tool_name=tool_name,
                            query=q,
                            code="BAD_TOOL_RETURN",
                            message=f"工具返回类型不是 dict，而是 {type(res).__name__}",
                            meta={"attempt": attempt, "timeout": timeout}
                        )
                    # 确保返回结构中有 meta，并注入执行元数据
                    res.setdefault("meta", {})
                    res["meta"].update(
                        {
                            "attempt": attempt,
                            "timeout": timeout,
                            "elapsed": round(time.time() - start, 3)
                        }
                    )
                    return res
                except Exception as e:
                    last_exc = e
                    # 非指定异常（如参数错误）：立即返回，不再重试
                    if _retry_on_tuple and not isinstance(e, _retry_on_tuple):
                        break

                # 需要重试则等待一段时间后继续
                if attempt < max_retries:
                    # 指数退避：延迟时间 = min(max_delay, base_delay * 2^attempt) * 随机偏移
                    # 随机偏移 (0.8 + 0.4*random) 在 [0.8, 1.2] 之间，避免雷群效应
                    delay = min(max_delay, base_delay * (2 ** attempt)) * (0.8 + 0.4 * random.random())
                    await asyncio.sleep(delay)

            # 全部失败：返回统一的失败结构，记录实际尝试次数
            return _err(
                tool_name=tool_name,
                query=q,
                code="TOOL_FAILED",
                message=f"工具调用失败（已重试）：{type(last_exc).__name__}: {last_exc}",
                meta={"retries": attempt, "timeout": timeout},
            )

        # 根据原函数是否异步，返回不同的包装函数
        if is_async:
            @functools.wraps(fn)
            async def wrapper(*args, **kwargs):
                """异步包装器：直接调用 _call_async"""
                return await _call_async(*args, **kwargs)
            return wrapper
        else:
            @functools.wraps(fn)
            def wrapper(*args, **kwargs):
                """
                同步包装器：需要处理两种场景：
                1. 无事件循环：直接用 asyncio.run() 启动新循环
                2. 已有事件循环（如 LangGraph/FastAPI）：不能用 asyncio.run()，改用线程池
                """
                # 兼容"已有事件循环"的运行环境（LangGraph / FastAPI 等）
                # asyncio.run() 在已运行的事件循环中会抛 RuntimeError，须改为新线程
                try:
                    loop = asyncio.get_running_loop()
                except RuntimeError:
                    loop = None
                if loop is not None and loop.is_running():
                    # 当前线程已有事件循环：在新线程中启动事件循环运行 _call_async
                    import concurrent.futures
                    with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
                        return executor.submit(asyncio.run, _call_async(*args, **kwargs)).result()
                # 当前线程无事件循环：创建新循环运行 _call_async
                return asyncio.run(_call_async(*args, **kwargs))
            return wrapper

    return decorator


# ================================================================================
# 工具 1：Web 搜索（Tavily API）
# ================================================================================
# 安装依赖：pip install tavily-python
# 配置环境变量：
#   - TAVILY_API_KEY：Tavily API 密钥（必须）
#   - MAX_SEARCH_RESULTS：最多返回多少条搜索结果（可选，默认 5）

@tool(
    "search",
    description=(
            "使用 Tavily 在互联网上搜索最新/可信信息。"
            "输入应是简洁的检索关键词或短句（不要太长）。"
            "适用于：新闻、天气、排名、价格、政策、版本号、需要核实的事实。"
    )
)
@with_retry(tool_name="search", max_retries=2, timeout=15)
async def search(query: str) -> dict[str, Any]:
    """
    Web 搜索工具 — 通过 Tavily API 搜索互联网最新信息。

    功能描述：
      - 接收用户查询，通过 Tavily API 搜索互联网
      - 裁剪搜索结果以适应上下文大小
      - 返回前几条最相关的结果（标题、链接、摘要、发布时间）

    参数：
      query (str)：搜索关键词或短句
                   - 不能为空
                   - 最多 300 字符（超出会自动截断）
                   - 可以使用自然语言问题或关键词

    返回：
      {
        “ok”: True/False,
        “data”: {
          “results”: [                    # 搜索结果列表
            {
              “title”: str,               # 搜索结果标题
              “url”: str,                 # 链接
              “content”: str,             # 摘要（最多 800 字）
              “score”: float,             # 匹配分数（可能为 None）
              “published_date”: str       # 发布日期（可能为 None）
            },
            ...
          ],
          “answer”: str                   # Tavily 生成的简洁回答（可选）
        },
        “error”: {...}                    # 失败时的错误信息
      }

    执行逻辑：
      1. 输入清理：strip() 并检查是否为空
      2. 输入截断：超过 300 字符自动截断
      3. 依赖检查：延迟 import，缺少依赖返回友好错误
      4. 密钥检查：检查 TAVILY_API_KEY 环境变量
      5. 配置读取：从 runtime context 或 MAX_SEARCH_RESULTS 环境变量读取返回数
      6. API 调用：使用 TavilySearch.ainvoke() 执行搜索
      7. 结果裁剪：通过 _shrink_search_results() 压缩到适当大小
      8. 异常处理：临时错误（超时、连接）raise 给 with_retry，永久错误返回 _err()

    错误场景：
      - BAD_INPUT：query 为空
      - MISSING_DEPENDENCY：未安装 langchain-tavily
      - MISSING_API_KEY：缺少 TAVILY_API_KEY 环境变量
      - SEARCH_FAILED：API 调用出错（非临时错误）
      - 超时/连接错误：raise，由 with_retry 自动重试 2 次

    示例：
      result = await search(“Python 3.12 最新功能”)
      # 返回关于 Python 3.12 的最新搜索结果
    """
    tool_name = "search"
    q = (query or "").strip()

    # ════════════════════ 阶段 1：输入校验 ════════════════════
    if not q:
        return _err(
            tool_name=tool_name,
            query=q,
            code="BAD_INPUT",
            message="query 不能为空，请提供要搜索的关键词。"
        )
    if len(q) > 300:
        q = q[:300]

    # ════════════════════ 阶段 2：依赖检查 ════════════════════
    # 延迟 import 是为了：如果这个工具没有被使用，就不会强制安装依赖
    # 或者如果依赖缺失，整个 tools.py 模块也不会因此 import 失败
    try:
        from langchain_tavily import TavilySearch
    except ImportError:
        return _err(
            tool_name=tool_name,
            query=q,
            code="MISSING_DEPENDENCY",
            message="未安装依赖 langchain-tavily。请执行：pip install langchain-tavily",
        )

    # ════════════════════ 阶段 3：密钥检查 ════════════════════
    # Tavily API 需要密钥认证，提前检查能更快给出友好错误
    if not os.getenv("TAVILY_API_KEY", "").strip():
        return _err(
            tool_name=tool_name,
            query=q,
            code="MISSING_API_KEY",
            message="缺少环境变量 TAVILY_API_KEY，搜索工具不可用。请先设置该 Key。",
        )

    # ════════════════════ 阶段 4：执行搜索 ════════════════════
    try:
        # 读取配置：最多返回几条结果
        max_results = 5
        try:
            # 优先从 LangGraph runtime context 读取
            runtime = get_runtime(Context)
            max_results = int(getattr(runtime.context, "max_search_results", 5))
        except Exception:
            # fallback：从环境变量读取
            max_results = int(os.getenv("MAX_SEARCH_RESULTS", "5") or "5")

        # 初始化搜索器
        wrapped = TavilySearch(max_results=max_results)

        # 调用搜索 API（Tavily 支持两种调用签名）
        try:
            raw = await wrapped.ainvoke({"query": q})
        except Exception:
            # 如果字典形式失败，尝试直接传字符串
            raw = await wrapped.ainvoke(q)

        # 裁剪结果以适应上下文
        data = _shrink_search_results(raw, max_items=min(5, max_results))
        return _ok(
            tool_name=tool_name,
            query=q,
            data=data,
            meta={"max_results": max_results},
        )

    # ════════════════════ 阶段 5：异常处理 ════════════════════
    except (TimeoutError, asyncio.TimeoutError, ConnectionError, OSError):
        # 临时错误：raise 给 with_retry 装饰器，自动重试
        # 不能在这里 try-except 吞掉，否则 with_retry 无法捕获
        raise
    except Exception as e:
        # 永久错误（API 返回 4xx、解析失败等）：返回失败结果，不重试
        return _err(
            tool_name=tool_name,
            query=q,
            code="SEARCH_FAILED",
            message=f"搜索工具调用失败：{type(e).__name__}: {e}",
        )


# ================================================================================
# 工具 2：Excel 表格生成
# ================================================================================
# 安装依赖：pip install openpyxl
# 配置环境变量：
#   - EXCEL_OUTPUT_DIR：Excel 文件输出目录（默认 ./outputs）
#   - RAG_MAX_CONTENT_CHARS：内容字符数上限（默认 800）

class ExcelInput(BaseModel):
    """
    Excel 表格创建工具的输入参数 Schema（Pydantic 模型）。

    属性：
      filename (str)：输出文件名，不需要后缀 .xlsx
                      示例："report" → 保存为 report.xlsx 或 report_timestamp.xlsx
      headers (list[str])：表头列表
                           示例：["姓名", "部门", "薪资"]
                           不能为空
      rows (list[list])：数据行，每行是一个列表
                         示例：[["张三", "技术部", "15000"], ["李四", "销售部", "12000"]]
                         必须与 headers 列数相同
      sheet_name (str)：工作表名称，默认 "Sheet1"（不能超过 31 字符，Excel 限制）
      mode (str)：写入模式，默认 "timestamp"
                  - "timestamp"：添加时间戳前缀，避免覆盖（report → report_20260310_144200.xlsx）
                  - "overwrite"：覆盖已有文件（危险，会丢失旧数据）
                  - "append"：追加到已有文件的最后（要求表头一致，否则拒绝）
    """
    filename: str = Field(description="文件名，例如 'report'，不需要加 .xlsx 后缀")
    headers: list[str] = Field(description="表头列表，例如 ['姓名', '部门', '薪资']")
    rows: list[list] = Field(description="数据行，每行是一个列表，顺序与表头一致")
    sheet_name: str = Field(default="Sheet1", description="工作表名称")
    mode: str = Field(default="timestamp", description="写入模式：timestamp/overwrite/append")


async def _make_excel_impl(
        filename: str,
        headers: list[str],
        rows: list[list[Any]],
        sheet_name: str = "Sheet1",
        mode: str = "timestamp",     # "timestamp" | "overwrite" | "append"
        keep_backup: bool = False,   # append/overwrite 时是否额外留一份时间戳备份
) -> dict[str, Any]:
    """
    Excel 表格生成的核心实现。

    功能描述：
      - 创建或修改 Excel 工作簿
      - 支持三种模式：timestamp（带时间戳）、overwrite（覆盖）、append（追加）
      - 自动美化表头：粗体、居中、灰色背景
      - 自动调整列宽以适应内容
      - 冻结表头行便于滚动查看

    参数：
      filename (str)：不带后缀的文件名（会自动补 .xlsx）
      headers (list[str])：表头，不能为空
      rows (list[list[Any]])：数据行列表
      sheet_name (str)：工作表名称（默认 "Sheet1"，不能超过 31 字符）
      mode (str)：模式选择
                  - "timestamp"：添加时间戳前缀，新建文件，不覆盖旧数据
                  - "overwrite"：覆盖已有的同名文件，危险操作
                  - "append"：追加到已有文件，要求表头一致
      keep_backup (bool)：在 append/overwrite 时是否保存时间戳备份

    执行流程：
      (1) 输入校验：filename/headers/rows 不能为空或格式错误
      (2) Mode 校验：mode 必须是允许值之一
      (3) 依赖检查：openpyxl 库
      (4) 输出目录创建：EXCEL_OUTPUT_DIR 环境变量，默认 ./outputs
      (5) 模式分支：
          - timestamp 或新建：创建新 Workbook，写入表头和样式
          - append：打开已有文件，检查表头一致性，找到末行后追加
      (6) 追加数据行：遍历 rows，每行用 ws.append()
      (7) 列宽调整：扫描前 200 行，计算最大宽度，设置列宽
      (8) 可选备份：append/overwrite 时保存 .backup_timestamp 文件
      (9) 保存并返回：返回保存路径和元数据

    错误处理：
      - BAD_INPUT：参数缺失或无效
      - MISSING_DEPENDENCY：缺少 openpyxl
      - HEADER_MISMATCH：append 时表头不一致
      - EXCEL_PERMISSION_DENIED：文件被占用或无写权限
      - EXCEL_CREATE_FAILED：其他创建错误

    示例：
      result = await _make_excel_impl(
          filename="report",
          headers=["姓名", "部门"],
          rows=[["张三", "技术"], ["李四", "销售"]],
          mode="timestamp"
      )
      # 返回：
      # {
      #   "ok": True,
      #   "data": {
      #     "path": "/path/to/report_20260310_144200.xlsx",
      #     "base_path": "/path/to/report.xlsx",
      #     "sheet": "Sheet1",
      #     "row_count": 2,
      #     "mode": "timestamp"
      #   }
      # }
    """
    tool_name = "make_excel_table"
    fn = (filename or "").strip()

    # ════════════════════ 阶段 1：输入校验 ════════════════════
    if not fn:
        return _err(tool_name=tool_name, query="", code="BAD_INPUT", message="filename 不能为空")
    if not fn.lower().endswith(".xlsx"):
        fn += ".xlsx"

    if not headers or not isinstance(headers, list):
        return _err(tool_name=tool_name, query=fn, code="BAD_INPUT", message="headers 必须是非空 list[str]")
    if rows is None or not isinstance(rows, list):
        return _err(tool_name=tool_name, query=fn, code="BAD_INPUT", message="rows 必须是 list[list[Any]]")

    # ════════════════════ 阶段 2：Mode 校验 ════════════════════
    valid_modes = {"timestamp", "overwrite", "append"}
    if mode not in valid_modes:
        return _err(
            tool_name=tool_name, query=fn, code="BAD_INPUT",
            message=f"mode 参数无效，必须是 {sorted(valid_modes)} 之一，当前值：{mode!r}"
        )

    # ════════════════════ 阶段 3：依赖检查 ════════════════════
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return _err(
            tool_name=tool_name,
            query=fn,
            code="MISSING_DEPENDENCY",
            message="缺少依赖 openpyxl。请安装：pip install openpyxl"
        )

    # ════════════════════ 阶段 4：输出目录准备 ════════════════════
    out_dir = os.getenv("EXCEL_OUTPUT_DIR", "./outputs")
    out_dir = Path(out_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    base_path = out_dir / fn

    # ════════════════════ 阶段 5a：确定输出路径 ════════════════════
    # 根据 mode 决定是否添加时间戳前缀
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if mode == "timestamp":
        # 新建带时间戳的文件，避免覆盖已有文件
        out_path = out_dir / f"{Path(fn).stem}_{stamp}.xlsx"
    else:
        # 覆盖或追加到固定路径
        out_path = base_path

    # ════════════════════ 阶段 5b：创建或打开工作簿 ════════════════════
    wb = None
    ws = None

    if mode == "append" and base_path.exists():
        # ════ 追加模式：打开已有文件 ════
        wb = load_workbook(base_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

        # 检查表头一致性（防止列错位）
        existing_headers = [ws.cell(row=1, column=i).value for i in range(1, len(headers) + 1)]

        if any(h is None for h in existing_headers):
            # 表头行为空（可能是新表），原地写入表头
            # 注意：不能用 ws.append([...])，因为那会追加到末尾，而不是写入第一行
            for col_idx, h in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=str(h))
        elif [str(x) for x in existing_headers] != [str(h) for h in headers]:
            # 表头不一致，拒绝追加（保护数据）
            return _err(
                tool_name=tool_name,
                query=str(base_path),
                code="HEADER_MISMATCH",
                message="追加失败：目标文件表头与本次 headers 不一致（为防止列错位已拒绝写入）",
                meta={"existing": existing_headers, "incoming": headers},
            )
    else:
        # ════ 创建新表 (timestamp/overwrite) ════
        wb = Workbook()
        ws = wb.active
        ws.title = (sheet_name or "Sheet1")[:31]  # Excel 工作表名不超过 31 字符
        ws.append([str(h) for h in headers])

        # 美化表头：粗体、居中、灰色背景
        header_font = Font(bold=True)
        header_align = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill("solid", fgColor="EDEDED")  # 灰色
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = header_align
            cell.fill = header_fill

        # 冻结表头行，便于向下滚动时仍能看到列名
        ws.freeze_panes = "A2"

    # ════════════════════ 阶段 6：追加数据行 ════════════════════
    # 无论是新建还是追加，都在这里写入数据行
    for r in rows:
        ws.append(list(r))

    # ════════════════════ 阶段 7：自动调整列宽 ════════════════════
    # 扫描前 200 行内容，计算最大宽度，自动设置列宽（提升用户体验）
    max_width_cap = 60  # 列宽上限，防止过宽
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        # 初始宽度：表头的长度
        max_len = len(str(headers[col_idx - 1] or ""))
        # 扫描数据行（只扫前 200 行以避免性能问题）
        scan_n = min(len(rows), 200)
        for i in range(scan_n):
            try:
                v = rows[i][col_idx - 1]
            except Exception:
                v = ""
            max_len = max(max_len, len(str(v or "")))
        # 设置列宽：最多 60，最少 10，加 2 作为边距
        ws.column_dimensions[col_letter].width = min(max_width_cap, max(10, max_len + 2))

    # ════════════════════ 阶段 8：可选备份 ════════════════════
    # 在 append/overwrite 时可选地保存时间戳备份，便于数据恢复
    if keep_backup and base_path.exists() and mode in {"append", "overwrite"}:
        backup_path = out_dir / f"{Path(fn).stem}_backup_{stamp}.xlsx"
        # 用二进制复制，快速且安全
        backup_path.write_bytes(base_path.read_bytes())

    # ════════════════════ 阶段 9：保存文件并返回 ════════════════════
    try:
        wb.save(out_path)
        return _ok(
            tool_name=tool_name,
            query=fn,
            data={
                "path": str(out_path),          # 实际保存路径（可能带时间戳）
                "base_path": str(base_path),    # 基础路径（无时间戳）
                "sheet": ws.title,              # 工作表名
                "row_count": len(rows),         # 数据行数（不含表头）
                "mode": mode
            },
            meta={"output_dir": str(out_dir)},
        )
    except PermissionError as e:
        # 权限错误（文件被占用、无写权限）
        return _err(
            tool_name=tool_name,
            query=str(out_path),
            code="EXCEL_PERMISSION_DENIED",
            message=f"Excel 保存失败（权限/被占用）：{e}",
        )
    except Exception as e:
        # 其他错误（磁盘满、格式错误等）
        return _err(
            tool_name=tool_name,
            query=str(out_path),
            code="EXCEL_CREATE_FAILED",
            message=f"Excel 创建失败：{type(e).__name__}: {e}",
        )


# 用 LangChain StructuredTool 包装 _make_excel_impl
# StructuredTool 会根据 args_schema 生成工具的 JSON Schema，供 LLM 调用
make_excel_table = StructuredTool.from_function(
    coroutine=_make_excel_impl,   # 异步函数用 coroutine= 参数
    name="make_excel_table",
    description=(
        "创建和管理 Excel 表格文件。支持新建、覆盖、追加模式。"
        "输入包含文件名、表头列表、数据行列表、工作表名、写入模式。"
        "自动美化表头样式，调整列宽，冻结表头行。"
    ),
    args_schema=ExcelInput,
)


# ================================================================================
# 工具 3：RAG 内部知识库查询
# ================================================================================
# 安装依赖：
#   pip install langchain-chroma langchain-huggingface sentence-transformers
#   pip install rank_bm25 langchain-community
# 配置环境变量：
#   - CHROMA_DB_PATH：Chroma 向量数据库路径（默认 ./chroma_db）
#   - EMBEDDING_MODEL：Embedding 模型（默认 BAAI/bge-small-zh-v1.5）
#   - RERANKER_MODEL：Reranker 模型（默认 BAAI/bge-reranker-base）
#   - RERANKER_THRESHOLD：Reranker 分数阈值（默认 -5）
#   - RAG_MAX_CONTENT_CHARS：返回内容的最大字符数（默认 800）

# ════════════════════════════════════════════════════════════════════════════════
# 模块级单例缓存机制（解决 BGE 模型重复加载问题）
# ════════════════════════════════════════════════════════════════════════════════
# 注意：原版每次调用 query_internal_knowledge 都会重新加载 BGE 模型（约 1-2GB），
# 造成严重的性能问题和内存泄漏。修复方案：用模块级单例缓存，首次初始化后复用。

_retriever_instance = None              # 缓存 BM25 + 向量 双路检索器
_reranker_retriever_instance = None     # 缓存 BGE Reranker 精排器
_retriever_lock = threading.Lock()      # 线程安全锁：保护 _retriever_instance 初始化
_reranker_lock = threading.Lock()       # 线程安全锁：保护 _reranker_retriever_instance 初始化


def _get_retriever():
    """
    构建并缓存双路检索器（BM25 + 向量）。

    功能描述：
      - BM25：基于关键词频率的精确匹配检索，适合术语匹配、人名、特定短语
      - 向量：基于语义相似度的模糊检索，适合语义相关但措辞不同的查询
      - 两路合并去重：BM25 和向量各取 Top-5，合并后去重，最多返回 10 条

    实现细节：
      - 不依赖 LangChain 的 EnsembleRetriever（功能单一）
      - 手动合并逻辑，更灵活且可控
      - Embedding 模型：HuggingFaceEmbeddings（支持中文）
      - 向量存储：Chroma（持久化，支持 GPU 加速检索）
      - BM25 索引：从 Chroma 中加载所有文档后构建

    线程安全：
      - 使用双重检查锁定（Double-Checked Locking）
      - 首次调用：获取锁 → 二次检查 → 初始化 → 释放锁
      - 后续调用：跳过锁，直接返回实例（快速路径）

    返回：
      {
        "bm25": BM25Retriever,         # BM25 检索器（可能为 None 如果知识库为空）
        "vector": BaseRetriever         # 向量检索器
      }

    异常处理：
      - 缺少 langchain-chroma：ImportError
      - 缺少 HuggingFaceEmbeddings：ImportError
      - 缺少 BM25Retriever：ImportError
      - 知识库为空：降级为纯向量模式（bm25 = None）
    """
    global _retriever_instance

    # 快速路径：已初始化，无锁返回（高频调用场景下避免锁开销）
    if _retriever_instance is not None:
        return _retriever_instance

    # 获取锁，进入临界区
    with _retriever_lock:
        # 二次检查：防止并发条件下重复初始化
        # （另一个线程可能在我们等锁期间已完成初始化）
        if _retriever_instance is not None:
            return _retriever_instance

        # ════ 依赖导入 ════
        try:
            from langchain_chroma import Chroma
        except ImportError:
            raise ImportError("缺少依赖 langchain-chroma。请执行：pip install langchain-chroma")

        try:
            from langchain_huggingface import HuggingFaceEmbeddings
        except ImportError:
            raise ImportError("缺少依赖 langchain-huggingface。请执行：pip install langchain-huggingface sentence-transformers")

        try:
            from langchain_community.retrievers import BM25Retriever
        except ImportError:
            raise ImportError("缺少依赖。请执行：pip install rank_bm25 langchain-community")

        from langchain_core.documents import Document

        # ════ 初始化向量存储 ════
        chroma_dir = os.getenv("CHROMA_DB_PATH", "./chroma_db")
        model_name = os.getenv("EMBEDDING_MODEL", "BAAI/bge-small-zh-v1.5")

        embeddings = HuggingFaceEmbeddings(model_name=model_name)
        vectorstore = Chroma(
            persist_directory=chroma_dir,
            embedding_function=embeddings,
        )
        vector_retriever = vectorstore.as_retriever(search_kwargs={"k": 5})

        # ════ 初始化 BM25 检索器 ════
        # 从 Chroma 向量库中提取所有文档（包括文本和元数据）
        all_docs = vectorstore.get()
        if not all_docs or not all_docs.get("documents"):
            # 知识库为空（可能首次运行，还没有数据），降级为纯向量检索
            print("[RAG] ⚠️ 知识库为空，降级为纯向量检索（请先导入文档）")
            _retriever_instance = {"bm25": None, "vector": vector_retriever}
            return _retriever_instance

        # 将 Chroma 返回的文本和元数据重新封装为 Document 对象（BM25 需要）
        docs_for_bm25 = [
            Document(page_content=text, metadata=meta)
            for text, meta in zip(all_docs["documents"], all_docs["metadatas"])
        ]
        bm25_retriever = BM25Retriever.from_documents(docs_for_bm25)
        bm25_retriever.k = 5  # BM25 也取 Top-5

        print(f"[RAG] ✅ 双路检索器初始化完成（BM25 文档数：{len(docs_for_bm25)}）")

        # ════ 存储单例 ════
        _retriever_instance = {
            "bm25": bm25_retriever,
            "vector": vector_retriever,
        }
        return _retriever_instance


async def _dual_retrieve(q: str) -> list:
    """
    双路召回 + 智能去重。

    功能描述：
      - 并发执行 BM25 和向量检索
      - 合并两路结果
      - 通过页面内容前缀进行去重（避免同一段落被多次返回）
      - 返回最多 10 条候选文档给 Reranker 精排

    执行流程：
      1. 获取检索器实例（使用缓存的单例）
      2. 并发执行：
         - BM25 检索（如果知识库非空）
         - 向量检索
      3. 合并：BM25 结果 + 向量结果（顺序很重要，BM25 优先级更高）
      4. 去重：用文档开头 100 字符作为 key，保留首次出现的文档
      5. 截断：最多返回 10 条（避免 Reranker 计算量过大）

    去重策略：
      - 使用 `doc.page_content[:100]` 作为去重 key
      - 去重 key 较短（100 字符），意味着相似但不完全相同的文档可能被同时返回
      - 可在环境变量中配置调整

    参数：
      q (str)：查询字符串

    返回：
      list[Document]：合并去重后的文档列表，最多 10 条
    """
    retriever = _get_retriever()

    bm25 = retriever.get("bm25")
    vector = retriever.get("vector")

    # ════ 并发执行两路检索（利用 asyncio.gather 并发） ════
    if bm25 is not None:
        # 知识库非空：并发执行 BM25 和向量检索
        bm25_docs, vector_docs = await asyncio.gather(
            asyncio.to_thread(bm25.invoke, q),      # BM25 在线程中运行（阻塞 I/O）
            asyncio.to_thread(vector.invoke, q),    # 向量检索也在线程中运行
        )
    else:
        # 知识库为空（降级模式）：仅向量检索
        vector_docs = await asyncio.to_thread(vector.invoke, q)
        bm25_docs = []

    # ════ 新增：RRF 融合打分 ════
    k = 60  # RRF 常数，工业标准值，一般不需要调整
    rrf_scores: dict[str, float] = {}
    doc_map: dict[str, object] = {}  # key —> Document 对象

    def get_key(doc) -> str:
        # 优先用 chunk_id，兜底用前100字符（兼容旧数据）
        return doc.metadata.get("chunk_id") or doc.page_content[:100]

    for rank, doc in enumerate(bm25_docs):
        key = get_key(doc)
        rrf_scores[key] = rrf_scores.get(key, 0.0) + 1 / (k + rank + 1)
        doc_map[key] = doc

    for rank, doc in enumerate(vector_docs):
        key = get_key(doc)
        rrf_scores[key] = rrf_scores.get(key, 0.0) + 1 / (k + rank + 1)
        doc_map[key] = doc

    sorted_keys = sorted(rrf_scores, key=lambda ks: rrf_scores[ks], reverse=True)

    return [doc_map[k] for k in sorted_keys[:10]]


def _get_reranker():
    """
    构建并缓存 BGE-Reranker 精排器。

    功能描述：
      - Cross-Encoder 模型，对 (query, document) 对进行相关性评分
      - 给双路召回的 10 条候选文档重新排序
      - 只保留分数高于阈值的文档（默认阈值 -5）
      - 最后返回 Top-3 给 LLM

    模型选择：
      - BAAI/bge-reranker-base：轻量级（~130M），速度快，精度中等
      - BAAI/bge-reranker-large：大型（~700M），精度高，速度较慢
      - 中文友好，专为 BGE 优化的 embedding 配套

    线程安全：
      - 使用双重检查锁定
      - 首次初始化时加载模型到 GPU/CPU（约 5-10 秒）
      - 后续调用使用缓存，无加载延迟

    返回：
      HuggingFaceCrossEncoder：cross-encoder 模型实例

    异常处理：
      - 缺少 langchain-community：ImportError
      - 模型下载失败：网络错误（由 HF 自动重试）
    """
    global _reranker_retriever_instance

    # 快速路径：已初始化，无锁返回
    if _reranker_retriever_instance is not None:
        return _reranker_retriever_instance

    # 获取锁，进入临界区
    with _reranker_lock:
        # 二次检查
        if _reranker_retriever_instance is not None:
            return _reranker_retriever_instance

        # ════ 依赖导入 ════
        try:
            from langchain_community.cross_encoders import HuggingFaceCrossEncoder
        except ImportError:
            raise ImportError("缺少依赖。请执行：pip install langchain-community sentence-transformers")

        # ════ 初始化 Reranker 模型 ════
        reranker_model = HuggingFaceCrossEncoder(
            model_name=os.getenv("RERANKER_MODEL", "BAAI/bge-reranker-base")
        )
        print("[RAG] ✅ Reranker 初始化完成（BAAI/bge-reranker-base）")

        # ════ 存储单例 ════
        _reranker_retriever_instance = reranker_model
        return _reranker_retriever_instance


async def _rerank(q: str, docs: list, top_n: int = 3) -> list:
    """
    使用 Cross-Encoder 对候选文档进行精排。

    功能描述：
      - 对每条候选文档计算与查询的相关性分数
      - 按分数降序排列
      - 通过阈值过滤（低分数视为无关）
      - 返回 Top-N 最相关的文档

    执行流程：
      1. 处理空输入 → 返回空列表
      2. 加载 Reranker 模型（首次调用时初始化）
      3. 构造 (query, document) 对
      4. 批量评分：reranker.score(pairs)
      5. 按分数降序排列
      6. 应用阈值过滤：score > threshold 才算相关
      7. 取 Top-N 结果

    分数范围与阈值：
      - BGE-Reranker 的分数范围：-10 到 +10（通常）
      - 默认阈值：-5（可通过 RERANKER_THRESHOLD 环境变量调整）
      - 分数 > -5 的文档被认为有相关性
      - 全部低于阈值 → 返回空列表（表示知识库中没有相关内容）

    参数：
      q (str)：查询字符串
      docs (list[Document])：候选文档列表（通常来自 _dual_retrieve，最多 10 条）
      top_n (int)：最多返回几条文档（默认 3）

    返回：
      list[Document]：精排后的相关文档，最多 top_n 条，按相关性从高到低排列

    异常处理：
      - 如果 Reranker 失败（网络、显存不足等），降级返回原始顺序的前 top_n 条
      - 记录警告日志便于诊断

    关键特性：
      - 异步无阻塞：通过 asyncio.to_thread 在线程中运行 Reranker（CPU 绑定）
      - 弹性降级：失败时返回原始结果而不是完全失败
      - 可配置阈值：通过环境变量调整相关性判断标准
    """
    # 处理空输入
    if not docs:
        return []

    try:
        # ════ 加载 Reranker 模型 ════
        # _get_reranker() 首次调用会加载模型（约 5-10 秒的阻塞 I/O）
        # 必须放在 asyncio.to_thread 中，避免阻塞事件循环
        reranker = await asyncio.to_thread(_get_reranker)

        # ════ 构造 (query, document) 对 ════
        pairs = [(q, doc.page_content) for doc in docs]

        # ════ 批量评分 ════
        # reranker.score() 是 CPU 绑定操作，在线程中运行
        scores = await asyncio.to_thread(reranker.score, pairs)

        # ════ 按分数降序排列 ════
        scored = sorted(zip(scores, docs), key=lambda x: x[0], reverse=True)

        # ════ 应用阈值过滤 ════
        # BGE-Reranker 的分数范围大约在 -10 到 +10 之间
        # 默认阈值 -5 可通过 RERANKER_THRESHOLD 环境变量调整
        # 只有分数 > 阈值 的文档才被认为与查询相关
        threshold = float(os.getenv("RERANKER_THRESHOLD", "0.1"))

        filtered = [(score, doc) for score, doc in scored if score > threshold]

        # 全部低于阈值 → 知识库中没有相关内容
        if not filtered:
            return []

        # ════ 返回 Top-N ════
        return [doc for _, doc in filtered[:top_n]]

    except Exception as e:
        print(f"[RAG] ⚠️ Reranker 失败：{e}")

        return []  # 改成返回空列表，而不是 docs[:top_n]


@tool(
    "query_internal_knowledge",
    description=(
        "查询公司内部的私有知识库、SOP文档和历史技术方案。"
        "当用户询问未公开的公司业务、内部规定、系统架构时，必须优先使用此工具。"
        "输入应为明确的检索关键词或陈述句。"
    )
)
@with_retry(tool_name="query_internal_knowledge", max_retries=2, timeout=30)
async def query_internal_knowledge(query: str) -> dict:
    """
    RAG 内部知识库查询工具 — 检索私有文档、SOP、技术方案等。

    功能描述：
      - 双路检索（BM25 + 向量）：精确匹配 + 语义匹配
      - Reranker 精排：重新评分候选文档，只保留相关内容
      - 智能拒答：知识库中没有相关内容时返回 has_relevant_content=False
      - 内容裁剪：只返回前 800 字符内容，保持上下文紧凑

    执行流程：
      1. 输入校验：检索词不能为空
      2. 双路召回（_dual_retrieve）：
         - BM25：关键词精确匹配，取 Top-5
         - 向量：语义匹配，取 Top-5
         - 合并去重：最多 10 条候选
      3. 如果候选为空：立即返回 has_relevant_content=False
      4. Reranker 精排（_rerank）：
         - 对候选文档评分
         - 保留分数 > -5 的文档
         - 取 Top-3 结果
      5. 结果构造：裁剪内容，保留来源信息
      6. 异常处理：
         - 临时错误（超时、连接）raise 给 with_retry 自动重试
         - 永久错误（初始化失败）返回失败结果

    参数：
      query (str)：查询关键词或问题陈述
                   示例："L1 scheme 的误差阶是多少？"
                   示例："反常扩散现象的数学建模方式"

    返回：
      {
        "ok": True/False,
        "data": {
          "results": [
            {
              "content": str,       # 文档片段（最多 800 字）
              "source": str         # 来源文件名（如 "paper.pdf"）
            },
            ...                     # 最多 3 条
          ],
          "has_relevant_content": bool  # 是否找到相关内容（重要！）
        },
        "meta": {
          "retrieved_count": int,       # 最终返回的相关文档数
          "candidates_count": int,      # 双路召回的候选数
          "stage": str                  # 执行阶段标记
        }
      }

    关键特性：
      - 拒答机制：has_relevant_content=False 表示知识库中无答案，不应该瞎编
      - 可追踪性：meta 中包含候选数、最终数，便于诊断检索质量
      - 重试保护：超时自动重试 2 次（via with_retry 装饰器）
      - 降级容错：Reranker 失败时降级返回 BM25 结果

    环境变量：
      - CHROMA_DB_PATH：向量数据库路径（默认 ./chroma_db）
      - EMBEDDING_MODEL：Embedding 模型（默认 BAAI/bge-small-zh-v1.5）
      - RERANKER_MODEL：Reranker 模型（默认 BAAI/bge-reranker-base）
      - RERANKER_THRESHOLD：相关性阈值（默认 -5）
      - RAG_MAX_CONTENT_CHARS：返回内容最大字符数（默认 800）

    使用示例：
      result = await query_internal_knowledge("Python asyncio 最佳实践")
      if result["ok"]:
          if result["data"]["has_relevant_content"]:
              for doc in result["data"]["results"]:
                  print(f"来源：{doc['source']}")
                  print(f"内容：{doc['content']}")
          else:
              print("知识库中没有相关内容")
    """
    tool_name = "query_internal_knowledge"
    q = (query or "").strip()

    # ════ 输入校验 ════
    if not q:
        return _err(
            tool_name=tool_name,
            query=q,
            code="BAD_INPUT",
            message="检索词不能为空"
        )

    try:
        # ════ 第一步：双路召回（BM25 + 向量） ════
        candidates = await _dual_retrieve(q)

        # 候选为空 → 知识库中没有相关内容
        if not candidates:
            return _ok(
                tool_name=tool_name,
                query=q,
                data={"results": [], "has_relevant_content": False},
                meta={"retrieved_count": 0, "stage": "dual_retrieve"}
            )

        # ════ 第二步：Reranker 精排，取 Top-3 ════
        docs = await _rerank(q, candidates, top_n=3)
        # 如果 len(docs) > 0 但查询明显无关，说明阈值问题已确认

        # ════ 第三步：构造返回结果 ════
        max_chars = int(os.getenv("RAG_MAX_CONTENT_CHARS", "800"))
        results = [
            {
                "content": _trim_text(doc.page_content, max_chars),
                "source": doc.metadata.get("source", "unknown"),
            }
            for doc in docs
        ]

        return _ok(
            tool_name=tool_name,
            query=q,
            data={
                "results": results,
                "has_relevant_content": len(docs) > 0  # 是否有相关内容
            },
            meta={
                "retrieved_count": len(results),      # 最终返回数
                "candidates_count": len(candidates),  # 候选数
                "stage": "dual_retrieve + rerank"
            }
        )

    # ════ 异常处理 ════
    except (TimeoutError, asyncio.TimeoutError, ConnectionError, OSError):
        # 临时错误：raise 给 with_retry 自动重试
        raise
    except Exception as e:
        # 永久错误：返回失败结果
        return _err(
            tool_name=tool_name,
            query=q,
            code="RAG_SEARCH_FAILED",
            message=f"知识库检索失败: {type(e).__name__}: {e}"
        )


# ================================================================================
# 工具导出：供 graph.py 中的 ToolNode 或 bind_tools 使用
# ================================================================================
# 说明：
#   这三个工具会被注册到 LangGraph 的工具节点中，LLM 可以根据需要调用任何一个
#   - search：Web 搜索，获取最新互联网信息
#   - make_excel_table：Excel 表格生成，导出数据
#   - query_internal_knowledge：私有知识库查询，获取内部文档和技术方案
#
# 使用示例（在 graph.py 中）：
#   from react_agent.tools import TOOLS
#   tools_node = ToolNode(TOOLS)  # 将工具列表传给 ToolNode
#   graph.add_node("tools", tools_node)
#
#   # 或者使用 bind_tools：
#   model = ChatAnthropic(model="claude-3-5-sonnet-20241022")
#   model_with_tools = model.bind_tools(TOOLS)

TOOLS = [search, make_excel_table, query_internal_knowledge]



